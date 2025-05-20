from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog
import sys
from PyQt5.QtGui import QWheelEvent
from PyQt5 import QtGui
from PyQt5.QtWidgets import QWidget
from PyQt5.QtGui import QPixmap,QIcon
from PyQt5.QtCore import Qt, QTimer, PYQT_VERSION_STR, QEvent
from PyQt5 import QtWidgets, QtGui, QtCore
import queue
import cv2
import os
from os.path import expanduser
import openpyxl 
import pandas as pd
from datetime import datetime
import sys
from PyQt5.QtWidgets import QApplication
import numpy as np
import cv2
import matplotlib.pyplot as plt 
import time
import platform
from serial.tools import list_ports
os.environ["SDL_JOYSTICK_ALLOW_BACKGROUND_EVENTS"] = "1"
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
try:
    import EasyPySpin
except Exception:
    pass

from classes.tracker_class import VideoThread
from classes.gui_widgets import Ui_MainWindow

from classes.robot_class import Robot
from classes.cell_class import Cell
from classes.arduino_class import ArduinoHandler
from classes.joystick_class import Mac_Controller,Linux_Controller,Windows_Controller
from classes.simulation_class import HelmholtzSimulator
from classes.projection_class import AxisProjection
from old.acoustic_class import AcousticClass

from classes.record_class import RecordThread



class MainWindow(QtWidgets.QMainWindow):
    positionChanged = QtCore.pyqtSignal(QtCore.QPoint)

    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent=parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        

        
        
        #self.showMaximized()

        #resize some widgets to fit the screen better
        screen  = QtWidgets.QDesktopWidget().screenGeometry(-1)
        
        self.window_width = screen.width()
        self.window_height = screen.height()
        self.resize(self.window_width, self.window_height)
        self.display_width = self.window_width# self.ui.frameGeometry().width()

        self.displayheightratio = 0.79
        self.framesliderheightratio = 0.031
        self.textheightratio = .129
        self.tabheightratio = 0.925
        self.tabheightratio = 0.925
        
        self.aspectratio = 1041/801
        self.resize_widgets()

    
      
        #create folder in homerdiractory of user
        if "Windows" in platform.platform():
            home_dir = expanduser("D:")
            new_dir_name = "Tracking Data"
            desktop_path = os.path.join(home_dir, "Microrobots")
            self.new_dir_path = os.path.join(desktop_path, new_dir_name)
            if not os.path.exists(self.new_dir_path):
                os.makedirs(self.new_dir_path)
        else:
            #home_dir = expanduser("~")
            #new_dir_name = "Tracking Data"
            #desktop_path = os.path.join(home_dir, "Desktop")
            
            self.new_dir_path = "Data"#os.path.join(desktop_path, new_dir_name)
            if not os.path.exists(self.new_dir_path):
                os.makedirs(self.new_dir_path)



        self.zoom_x, self.zoom_y, self.zoomscale, self.scrollamount = 1,0,0,0
        self.croppedresult = None
        self.frame_number = 0
        self.robots = []
        self.cells = []
        self.videopath = 0
        self.cap = None
        self.tracker = None
        self.populate_serial_ports()
        self.arduino = None

        #record variables
        self.recorder = None
        self.frame_queue = queue.Queue(maxsize=100)  # make a queue to store frames in for the recording feature
        self.output_file_name = "output"

        self.save_status = False
        self.output_workbook = None
        self.ricochet_counter_x = [0]
        self.ricochet_counter_y = [0]
        
        
        self.drawing = False
        self.acoustic_frequency = 0
        self.gradient_status = 0
        self.equal_field_status = 0
        self.magnetic_field_list = []
        
        self.actions = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.Bx, self.By, self.Bz = 0,0,0
        self.Mx, self.My, self.Mz = 0,0,0
        self.alpha, self.gamma, self.psi, self.freq = 0,0,0,0
        self.field_magnitude = 100

      
        #control tab functions
        self.control_status = False
        self.joystick_status = False
        self.manual_status = False


  
        if "mac" in platform.platform():
            self.tbprint("Detected OS: macos")
            self.controller_actions = Mac_Controller()
        elif "Linux" in platform.platform():
            self.tbprint("Detected OS: Linux")
            self.controller_actions = Linux_Controller()
        elif "Windows" in platform.platform():
            self.tbprint("Detected OS:  Windows")
            self.controller_actions = Windows_Controller()
        else:
            self.tbprint("undetected operating system")
        

        
        
        #define, simulator class, pojection class, and acoustic class
        self.simulator = HelmholtzSimulator(self.ui.magneticfieldsimlabel, width=310, height=310, dpi=200)
        self.projection = AxisProjection()
        self.acoustic_module = AcousticClass()
        
        

        self.setFile()
        
        pygame.init()
        if pygame.joystick.get_count() == 0:
            self.tbprint("No Joystick Connected...")
            
        else:
            self.joystick = pygame.joystick.Joystick(0)
            self.joystick.init()
            self.tbprint("Connected to: "+str(self.joystick.get_name()))
        
      
        self.sensorupdatetimer = QTimer(self)
        self.sensorupdatetimer.timeout.connect(self.update_sensor_label)
        self.sensorupdatetimer.start(25)  # Update every 500 ms
        self.bx_sensor = 0
        self.by_sensor = 0 
        self.bz_sensor = 0

     

        #tracker tab functions
        self.ui.pausebutton.hide()
        self.ui.leftbutton.hide()
        self.ui.rightbutton.hide()
        self.ui.choosevideobutton.clicked.connect(self.selectFile)
        self.ui.startbutton.clicked.connect(self.start)
        self.ui.pausebutton.clicked.connect(self.pause)
        self.ui.rightbutton.clicked.connect(self.frameright)
        self.ui.leftbutton.clicked.connect(self.frameleft)
        self.ui.maskbutton.clicked.connect(self.showmask)
        self.ui.maskinvert_checkBox.toggled.connect(self.invertmaskcommand)
        self.ui.robotmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotcroplengthbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellcroplengthbox.valueChanged.connect(self.get_slider_vals)
        self.ui.gradient_status_checkbox.toggled.connect(self.gradientcommand)
        self.ui.equal_field_checkbox.toggled.connect(self.equalfieldcommand)
        self.ui.savedatabutton.clicked.connect(self.savedata)
        self.ui.VideoFeedLabel.installEventFilter(self)
        self.ui.recordbutton.clicked.connect(self.toggle_recording)
        self.ui.controlbutton.clicked.connect(self.toggle_control_status)
        self.ui.memorybox.valueChanged.connect(self.get_slider_vals)
        self.ui.RRTtreesizebox.valueChanged.connect(self.get_slider_vals)
        self.ui.arrivalthreshbox.valueChanged.connect(self.get_slider_vals)
        self.ui.magneticfrequencydial.valueChanged.connect(self.get_slider_vals)
        self.ui.gammadial.valueChanged.connect(self.get_slider_vals)
        self.ui.psidial.valueChanged.connect(self.get_slider_vals)
        self.ui.applyacousticbutton.clicked.connect(self.apply_acoustic)
        self.ui.acousticfreq_spinBox.valueChanged.connect(self.get_acoustic_frequency)
        self.ui.alphaspinBox.valueChanged.connect(self.spinbox_alphachanged)
        self.ui.alphadial.valueChanged.connect(self.dial_alphachanged)
        self.ui.resetdefaultbutton.clicked.connect(self.resetparams)
        self.ui.simulationbutton.clicked.connect(self.toggle_simulation)
        self.ui.orientradio.toggled.connect(self.checkorient)
        self.ui.pushradio.toggled.connect(self.checkpush)
        self.ui.objectivebox.valueChanged.connect(self.get_objective)
        self.ui.exposurebox.valueChanged.connect(self.get_exposure)
        self.ui.joystickbutton.clicked.connect(self.toggle_joystick_status)
        self.ui.autoacousticbutton.clicked.connect(self.toggle_autoacoustic)
        self.ui.manualapplybutton.clicked.connect(self.get_manual_bfieldbuttons)
        self.ui.manualfieldBx.valueChanged.connect(self.get_slider_vals)
        self.ui.manualfieldBy.valueChanged.connect(self.get_slider_vals)
        self.ui.manualfieldBz.valueChanged.connect(self.get_slider_vals)
        self.ui.croppedmasktoggle.clicked.connect(self.showcroppedoriginal)
        self.ui.croppedrecordbutton.clicked.connect(self.croppedrecordfunction)
        self.ui.import_excel_actions.clicked.connect(self.read_excel_actions)
        self.ui.apply_actions.clicked.connect(self.apply_excel_actions)
        self.ui.arduino_portbox.currentTextChanged.connect(self.handle_port_change)

        self.ui.cleartrackingbutton.clicked.connect(self.clear_tracking)

        self.excel_file_name = None
        self.excel_actions_df = None
        self.excel_actions_status = False


        self.ui.makeshapebutton.clicked.connect(self.makeshape_trajectory)

        

    def clear_tracking(self):
        if self.tracker is not None:
            del self.tracker.robot_list[:]
            del self.tracker.cell_list[:]
            del self.magnetic_field_list[:]
            del self.robots[:]
            del self.cells[:]
            self.apply_actions(False)
    



        
    def makeshape_trajectory(self):
        if self.tracker is not None:
            if len(self.tracker.robot_list)>0:
                node_number = self.ui.shapemaker_nodes.value()

                center_x = self.video_width // 2
                center_y = self.video_height // 2
                radius = min(self.video_width, self.video_height) // 4  # Assume circle fits within a quarter of the image
                
                coordinates = []
                for i in range(node_number):
                    theta = 2 * np.pi * i / node_number
                    x = center_x + int(radius * np.cos(theta))
                    y = center_y + int(radius * np.sin(theta))
                    coordinates.append((x, y))
                
                coordinates.append(coordinates[0])
                
                self.tracker.robot_list[-1].trajectory = coordinates 
        
    

    def update_sensor_label(self):
        # Replace this with your actual value source
        self.ui.bxlabel.setText("Bx: "+str(self.bx_sensor))
        self.ui.bylabel.setText("By: "+str(self.by_sensor))
        self.ui.bzlabel.setText("Bz: "+str(self.bz_sensor))

    

    def update_actions(self, actions, stopped, robot_list, cell_list):


        #read hall effect sensor data from arduino
        sensor = self.arduino.receive()
        self.bx_sensor = -1 * round(sensor[0], 1)   #Bx sensor sign is switched
        self.by_sensor = round(sensor[1], 1)
        self.bz_sensor = round(sensor[2], 1)


   


        self.frame_number+=1
        #toggle between alpha and orient
        

        #output actions if control status is on
        if self.ui.autoacousticbutton.isChecked():
            self.acoustic_frequency = actions[-1]   
        
        if self.control_status == True:

            if not self.ui.pushradio.isChecked():
                self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, _  = actions    
            
                
                self.gamma = np.radians(self.ui.gammadial.value())
                self.psi = np.radians(self.ui.psidial.value())
                if self.ui.orientradio.isChecked():
                    self.freq = 0
                else:
                    self.freq = self.ui.magneticfrequencydial.value()

                if stopped == True:
                    self.apply_actions(False)
            else:
                print("pushing")

            
        #if joystick is on use the joystick though
        elif self.joystick_status == True:
            type, self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, _ = self.controller_actions.run(self.joystick)
            self.psi = np.radians(self.ui.psidial.value())
            
            if type == 1:
                self.gamma = np.radians(180)
                self.freq = self.ui.magneticfrequencydial.value()
            
            elif type == 2:
                self.gamma = np.radians(0)
                self.freq = self.ui.magneticfrequencydial.value()
            
            else:
                self.gamma = np.radians(self.ui.gammadial.value())
                if self.freq != 0:
                    self.freq = self.ui.magneticfrequencydial.value()
                    
                        
        
        
        elif self.manual_status == True:
            self.Bx = self.ui.manualfieldBx.value()/100
            self.By = self.ui.manualfieldBy.value()/100
            self.Bz = self.ui.manualfieldBz.value()/100
            self.freq = self.ui.magneticfrequencydial.value()
            self.gamma = np.radians(self.ui.gammadial.value())
            self.psi = np.radians(self.ui.psidial.value())
            self.alpha = np.radians(self.ui.alphadial.value())
            
            #ricochet conditions, too close to the x or y borders flip the conditions
            if self.ui.ricochet_effect_checkbox.isChecked():
                if len(self.tracker.robot_list) > 0:
                    for i in range(len(self.tracker.robot_list)):
                        bot = self.tracker.robot_list[i]
                        bot_pos_x = int(bot.position_list[-1][0])
                        bot_pos_y = int(bot.position_list[-1][1])
                        
                        vx = bot.velocity_list[-1][0]
                        vy = bot.velocity_list[-1][1] 
                        boundary = 200
                        #ricochet conditions, too close to the x or y borders
                        
                        if (bot_pos_x <= boundary and vx < 0) and (self.frame_number - self.ricochet_counter_x[-1] > 30):
                            vx = -vx
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_x.append(self.frame_number)

                        if  (bot_pos_x >= self.video_width - boundary and vx > 0) and (self.frame_number - self.ricochet_counter_x[-1] > 30):   
                            vx = -vx
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_x.append(self.frame_number)                  
                        
                        if (bot_pos_y <= boundary and vy < 0) and (self.frame_number - self.ricochet_counter_y[-1] > 30):
                            vy = -vy
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_y.append(self.frame_number)
            
                        if (bot_pos_y >= self.video_height - boundary and vy > 0) and (self.frame_number - self.ricochet_counter_y[-1] > 30):# and (self.frame_number - self.ricochet_counter_y[-1] > 30): #if the bot hits the top wall   
                            vy = -vy
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_y.append(self.frame_number)
                           
                    
                    
             
                        
            


        elif self.excel_actions_status == True and self.excel_actions_df is not None:            
            self.actions_counter +=1
            if self.actions_counter < self.excel_actions_df['Frame'].iloc[-1]:
                filtered_row = self.excel_actions_df[self.excel_actions_df['Frame'] == self.actions_counter]
                
                self.Bx = float(filtered_row["Bx"])
                self.By = float(filtered_row["By"])
                self.Bz = float(filtered_row["Bz"])
                self.alpha = float(filtered_row["Alpha"])
                self.gamma = float(filtered_row["Gamma"])
                self.freq = float(filtered_row["Rolling Frequency"])
                self.psi = float(filtered_row["Psi"])
                self.gradient = float(filtered_row["Gradient?"])
                self.equal_field_status = float(filtered_row["Equal Field?"])
                self.acoustic_freq = float(filtered_row["Acoustic Frequency"])
            
            else:
                self.excel_actions_status = False
                self.ui.apply_actions.setText("Apply")
                self.ui.apply_actions.setChecked(False)
                self.apply_actions(False)
            
        

    
        #DEFINE CURRENT ROBOT PARAMS TO A LIST
        if len(robot_list) > 0:
            self.robots = []
            for bot in robot_list:
                currentbot_params = [bot.frame_list[-1],
                                     bot.times[-1],
                                     bot.position_list[-1][0]* self.tracker.pixel2um,
                                     bot.position_list[-1][1]* self.tracker.pixel2um, 
                                     bot.velocity_list[-1][0]* self.tracker.pixel2um, 
                                     bot.velocity_list[-1][1]* self.tracker.pixel2um,
                                     bot.velocity_list[-1][2]* self.tracker.pixel2um,
                                     bot.blur_list[-1],
                                     bot.area_list[-1]* (self.tracker.pixel2um**2),
                                     bot.pixel2um,
                                     [[x * self.tracker.pixel2um, y * self.tracker.pixel2um] for x, y in bot.trajectory]
                                    ]
                
                self.robots.append(currentbot_params)
           
      
        #DEFINE CURRENT CELL PARAMS TO A LIST
        if len(cell_list) > 0:
            self.cells = []
            for cell in cell_list:
                currentcell_params = [cell.frame_list[-1],
                                     cell.times[-1],
                                     cell.position_list[-1][0]* self.tracker.pixel2um,
                                     cell.position_list[-1][1]* self.tracker.pixel2um, 
                                     cell.velocity_list[-1][0]* self.tracker.pixel2um, 
                                     cell.velocity_list[-1][1]* self.tracker.pixel2um,
                                     cell.velocity_list[-1][2]* self.tracker.pixel2um,
                                     cell.blur_list[-1],
                                     cell.area_list[-1]* (self.tracker.pixel2um**2),
                                     cell.pixel2um
                                    ]
                
                self.cells.append(currentcell_params)
        
        #DEFINE CURRENT MAGNETIC FIELD OUTPUT TO A LIST 
        
        self.actions = [self.frame_number, self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.gradient_status,self.equal_field_status,
                        self.acoustic_frequency, self.bx_sensor, self.by_sensor, self.bz_sensor] 
       
        self.magnetic_field_list.append(self.actions)
        self.apply_actions(True)
        
        

        #IF SAVE STATUS THEN CONTINOUSLY SAVE THE CURRENT ROBOT PARAMS AND MAGNETIC FIELD PARAMS TO AN EXCEL ROWS
        if self.save_status == True:
            self.magnetic_field_sheet.append(self.actions)
            for (sheet, bot) in zip(self.robot_params_sheets,self.robots):
                sheet.append(bot[:-1])
            for (sheet, cell) in zip(self.cell_params_sheets,self.cells):
                sheet.append(cell[:-1])


        





    def apply_actions(self, status):
        #the purpose of this function is to output the actions via arduino, 
        # show the actions via the simulator
        # and record the actions by appending the field_list
        if self.freq > 0:
            if self.ui.swimradio.isChecked():
                self.simulator.roll = False
            elif self.ui.rollradio.isChecked():
                self.alpha = self.alpha + np.pi/2
                self.simulator.roll = True

        #zero output
        if status == False:
            self.manual_status = False
            self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.acoustic_frequency = 0,0,0,0,0,0,0,0

        #output current actions to simulator

        self.simulator.Bx = self.Bx
        self.simulator.By = self.By
        self.simulator.Bz = self.Bz
        self.simulator.alpha = self.alpha
        self.simulator.gamma = self.gamma
        self.simulator.psi = self.psi
        self.simulator.freq = self.freq
        self.simulator.omega = 2 * np.pi * self.simulator.freq

        #send arduino commands
        self.arduino.send(self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.gradient_status, self.equal_field_status, self.acoustic_frequency)


    def toggle_recording(self):
        if self.cap is not None:
            if self.recorder and self.recorder.running:
                self.recorder.stop()
                self.recorder = None
                self.stop_data_record()
                self.ui.recordbutton.setText("Record Video")
                self.tbprint("Recording Stopped")
            else:
                self.frame_queue.queue.clear()
                 # set filename
                self.output_file_name = self.ui.videoNameLineEdit.text().strip()
                if not self.output_file_name:
                    self.output_file_name = "output"

                file_path  = os.path.join(self.new_dir_path, self.output_file_name+".mp4")
                self.recorder = RecordThread(self.frame_queue, file_path, self.videofps)
                self.recorder.start()
                self.start_data_record()
                self.ui.recordbutton.setText("Stop Record")
                self.tbprint("Recording Video...")



    def start_data_record(self):
        self.frame_number = 0
        self.tracker.framenum = 1
        self.tracker.start_time = time.time()
        self.tracker.time_stamp = 0
        
        self.output_workbook = openpyxl.Workbook()
            
        #create sheet for magneti field actions
        self.magnetic_field_sheet = self.output_workbook.create_sheet(title="Magnetic Field Actions")#self.output_workbook.active
        self.magnetic_field_sheet.append(["Frame","Bx", "By", "Bz", "Alpha", "Gamma", "Rolling Frequency", "Psi", "Gradient?","Equal Field?", "Acoustic Frequency","Sensor Bx", "Sensor By", "Sensor Bz"])

        #create sheet for robot data
        self.robot_params_sheets = []
        for i in range(len(self.robots)):
            robot_sheet = self.output_workbook.create_sheet(title= "Robot {}".format(i+1))
            robot_sheet.append(["Frame","Time(s)","Pos X (um)", "Pos Y (um)", "Vel X (um/s)", "Vel Y (um/s)", "Vel Mag (um/s)", "Blur", "Area (um^2)","pixel2um","Path X (um)", "Path Y (um)"])
            self.robot_params_sheets.append(robot_sheet)
        
        #create sheet for robot data
        self.cell_params_sheets = []
        for i in range(len(self.cells)):
            cell_sheet = self.output_workbook.create_sheet(title= "Cell {}".format(i+1))
            cell_sheet.append(["Frame","Time(s)","Pos X (um)", "Pos Y (um)", "Vel X (um/s)", "Vel Y (um/s)", "Vel Mag (um/s)", "Blur", "Area (um^2)","pixel2um"])
            self.cell_params_sheets.append(cell_sheet)

        #tell update_actions function to start appending data to the sheets
        self.save_status = True




    




    def stop_data_record(self):
        #tell update_actions function to stop appending data to the sheets
        
        self.save_status = False
        file_path  = os.path.join(self.new_dir_path, self.output_file_name+".xlsx")
        
        #add trajectory to file after the fact
        if self.output_workbook is not None:
            if len((self.robot_params_sheets)) > 0:
                try:
                    for i in range(len((self.robot_params_sheets))):
                        for idx,(x,y) in enumerate(self.robots[i][-1]):
                            self.robot_params_sheets[i].cell(row=idx+2, column=16).value = x
                            self.robot_params_sheets[i].cell(row=idx+2, column=17).value = y
                except Exception:
                    pass
                try:
                    for i in range(len((self.robot_params_sheets))):
                        for idx,(x,y) in enumerate(self.robots[i][-1]):
                            self.robot_params_sheets[i].cell(row=idx+2, column=16).value = x
                            self.robot_params_sheets[i].cell(row=idx+2, column=17).value = y
                except Exception:
                    pass
            #save and close workbook
            self.output_workbook.remove(self.output_workbook["Sheet"])
            self.output_workbook.save(file_path)

            self.output_workbook.close()
            self.output_workbook = None

    
    def savedata(self):
        if self.ui.savedatabutton.isChecked():
            self.ui.savedatabutton.setText("Stop")
            self.start_data_record()
            self.output_file_name = self.ui.videoNameLineEdit.text().strip()
            if not self.output_file_name:
                self.output_file_name = "output"
        else:
            self.ui.savedatabutton.setText("Save Data")
            self.stop_data_record()
            



    def read_excel_actions(self):
        options = QFileDialog.Options()
        self.excel_file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)", options=options)
        if self.excel_file_name:
            self.excel_actions_df = pd.read_excel(self.excel_file_name)
            
        
    def apply_excel_actions(self):
        if self.ui.apply_actions.isChecked():
            self.excel_actions_status = True
            self.actions_counter = 0
            self.ui.apply_actions.setText("Stop")
        else:
            self.excel_actions_status = False
            self.ui.apply_actions.setText("Apply")
            self.apply_actions(False)
    
    


    
    def toggle_simulation(self):
        if self.ui.simulationbutton.isChecked():
            self.simulator.start()
            self.tbprint("Simulation Off")
            self.ui.simulationbutton.setText("Simulation Off")
        else:
            self.simulator.stop()
            self.tbprint("Simulation On")
            self.ui.simulationbutton.setText("Simulation On")
   
    
    
    def toggle_control_status(self): 
        if self.ui.controlbutton.isChecked():
            self.control_status = True
            self.ui.controlbutton.setText("Stop")
            self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
        else:
            self.control_status = False
            self.ui.controlbutton.setText("Control")
            self.tbprint("Control Off")
            self.apply_actions(False)
            
            
    

    def toggle_joystick_status(self):
        if pygame.joystick.get_count() != 0:
            if self.ui.joystickbutton.isChecked():
                self.joystick_status = True
                self.ui.joystickbutton.setText("Stop")
                self.tbprint("Joystick On")
            else:
                self.joystick_status = False
                self.ui.joystickbutton.setText("Joystick")
                self.tbprint("Joystick Off")
                self.apply_actions(False)
        else:
            self.tbprint("No Joystick Connected...")

    def toggle_autoacoustic(self):
        if self.tracker is not None:
            if self.ui.autoacousticbutton.isChecked():
                self.tracker.autoacousticstatus = True
                self.ui.led.setStyleSheet("\n"
"                background-color: rgb(0, 255, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(0, 255, 0);\n"
"         \n"
"                padding: 6px;")
            else:
                self.tracker.autoacousticstatus = False
                self.acoustic_frequency = 0
                self.ui.led.setStyleSheet("\n"
"                background-color: rgb(255, 0, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(255, 0, 0);\n"
"         \n"
"                padding: 6px;")
                

                

    def get_acoustic_frequency(self):
        if self.ui.applyacousticbutton.isChecked():
            self.acoustic_frequency = self.ui.acousticfreq_spinBox.value()
            #self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
            self.apply_acoustic()
        
    
    def apply_acoustic(self):
        if self.ui.applyacousticbutton.isChecked():
            self.ui.applyacousticbutton.setText("Stop")
            #self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
            self.acoustic_frequency = self.ui.acousticfreq_spinBox.value()
            #self.acoustic_module.start(self.acoustic_frequency, 0)
            #self.apply_actions(True)
            self.ui.led.setStyleSheet("\n"
"                background-color: rgb(0, 255, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(0, 255, 0);\n"
"         \n"
"                padding: 6px;")
        
        else:
            self.ui.applyacousticbutton.setText("Apply")
            #self.tbprint("Acoustic Module Off")
            #self.acoustic_module.stop()
            self.acoustic_frequency = 0
            self.ui.led.setStyleSheet("\n"
"                background-color: rgb(255, 0, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(255, 0, 0);\n"
"         \n"
"                padding: 6px;")
            #self.apply_actions(False)
       
        

    def tbprint(self, text):
        #print to textbox
        self.ui.plainTextEdit.appendPlainText("$ "+ text)
    

    def convert_coords(self,pos):
        #need a way to convert the video position of mouse to the actually coordinate in the window
        newx = int(pos.x() * (self.video_width / self.display_width)) 
        newy = int(pos.y() * (self.video_height / self.display_height))
        return newx, newy
    
    
    def keyPressEvent(self, event):
        # Check if the event type is QEvent.KeyPress
        if event.type() == QEvent.KeyPress:
            # Check if the key pressed is 'c'
            if event.key() == Qt.Key_C:
                print('c')
        # Call the base class method for default processing
        return super().keyPressEvent(event)
 

    

    def eventFilter(self, object, event):
        
        if object is self.ui.VideoFeedLabel: 
            if self.tracker is not None:
                
                    
                if event.type() == QtCore.QEvent.MouseButtonPress:   
                    if event.buttons() == QtCore.Qt.LeftButton:
                        newx, newy = self.convert_coords(event.pos())
                        #generate original bounding box
                        
                        #reset algorithm nodes
                        self.tracker.control_robot.reset()

                        if self.ui.robotmask_radio.isChecked():
                            x_1 = int(newx - self.ui.robotcroplengthbox.value()  / 2)
                            y_1 = int(newy - self.ui.robotcroplengthbox.value()  / 2)
                            w = self.ui.robotcroplengthbox.value()
                            h = self.ui.robotcroplengthbox.value()

                            robot = Robot()  # create robot instance
                            robot.add_frame(self.frame_number)
                            robot.add_time(0)
                            robot.add_position([newx,newy])
                            robot.add_velocity([0,0,0])
                            robot.add_crop([x_1, y_1, w, h])
                            robot.add_area(0)
                            robot.add_blur(0)
                        
                            robot.crop_length = self.ui.robotcroplengthbox.value()
                            self.tracker.robot_list.append(robot) #this has to include tracker.robot_list because I need to add it to that class
                        
                        elif self.ui.cellmask_radio.isChecked():
                            x_1 = int(newx - self.ui.cellcroplengthbox.value()  / 2)
                            y_1 = int(newy - self.ui.cellcroplengthbox.value()  / 2)
                            w = self.ui.cellcroplengthbox.value()
                            h = self.ui.cellcroplengthbox.value()

                            cell = Cell()  # create robot instance
                            cell.add_frame(self.frame_number)
                            cell.add_time(0)
                            cell.add_position([newx,newy])
                            cell.add_velocity([0,0,0])
                            cell.add_crop([x_1, y_1, w, h])
                            cell.add_area(0)
                            cell.add_blur(0)
                           
                            cell.crop_length = self.ui.cellcroplengthbox.value()
                            
                            self.tracker.cell_list.append(cell) #this has to include tracker.robot_list because I need to add it to that class

               
                    
                    
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = True
                        newx, newy = self.convert_coords(event.pos())
                        if len(self.tracker.robot_list) > 0:
                            self.tracker.control_robot.reset()
                            self.tracker.control_robot.arrived = False
                            self.tracker.robot_list[-1].add_trajectory([newx, newy])
                
                
                    if event.buttons() == QtCore.Qt.MiddleButton: 
                        del self.tracker.robot_list[:]
                        del self.tracker.cell_list[:]
                        del self.magnetic_field_list[:]
                        del self.robots[:]
                        del self.cells[:]
                        self.apply_actions(False)
                       
                    
                            
                elif event.type() == QtCore.QEvent.MouseMove:
                    self.zoom_x, self.zoom_y = self.convert_coords(event.pos())

                    if event.buttons() == QtCore.Qt.RightButton:
                        if self.drawing == True:
                            if len(self.tracker.robot_list)>0:
                                newx, newy = self.convert_coords(event.pos())
                                
                                self.tracker.robot_list[-1].add_trajectory([newx, newy])
                                
                
                elif event.type() == QtCore.QEvent.MouseButtonRelease:
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = False
                        
                if event.type() ==  QtCore.QEvent.Wheel:
                    steps = event.angleDelta().y() 
                    
                    self.scrollamount += (steps and steps / abs(steps/0.5))
                    self.scrollamount = max(min(self.scrollamount,20.0),1.0)
                    self.zoomscale = self.scrollamount

        
        return super().eventFilter(object, event)
            
            

    def update_image(self, frame):
        """Updates the image_label with a new opencv image"""
        #display projection
        if self.ui.toggledisplayvisualscheckbox.isChecked():
            if self.control_status == True or self.joystick_status == True or self.manual_status == True or self.excel_actions_status == True :
                self.projection.roll = self.ui.rollradio.isChecked()
                self.projection.gradient = self.gradient_status


                frame, self.projection.draw_sideview(frame,self.Bx,self.By,self.Bz,self.alpha,self.gamma,self.video_width,self.video_height)
                frame, self.projection.draw_topview(frame,self.Bx,self.By,self.Bz,self.alpha,self.gamma,self.video_width,self.video_height)
                

                rotatingfield = "alpha: {:.0f}, gamma: {:.0f}, psi: {:.0f}, freq: {:.0f}".format(np.degrees(self.alpha), np.degrees(self.gamma), np.degrees(self.psi), self.freq) #adding 90 to alpha for display purposes only
                
                cv2.putText(frame, rotatingfield,
                    (int(self.video_width / 1.8),int(self.video_height / 20)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    fontScale=1.5, 
                    thickness=5,
                    color = (0, 0, 0),
                )
            
            acousticfreq = f'{self.acoustic_frequency:,} Hz'
            cv2.putText(frame, acousticfreq,
                (int(self.video_width / 8),int(self.video_height / 14)),
                cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=1.5, 
                thickness=5,
                color = (0, 0, 0),
            )

            
        

        
        
        frame = self.handle_zoom(frame)
    

        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
      
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(self.display_width, self.display_height, Qt.KeepAspectRatio)
        qt_img = QPixmap.fromImage(p)
       
        #update frame slider too
        self.ui.framelabel.setText("Frame:"+str(self.frame_number))
        if self.videopath !=0:
            self.ui.frameslider.setValue(self.tracker.framenum)
        
        #also update robot info
        if len(self.robots) > 0:
            area = self.robots[-1][8] 
            robot_diameter = round(np.sqrt(4*area/np.pi),1)
            self.ui.vellcdnum.display(int(self.robots[-1][6]))
            self.ui.blurlcdnum.display(int(self.robots[-1][7]))
            self.ui.sizelcdnum.display(robot_diameter)
                
       
        self.ui.VideoFeedLabel.setPixmap(qt_img)


        #add frame to recording queue if record button is pressed
        if self.recorder and self.recorder.running:
            try:
                self.frame_queue.put_nowait(frame)
            except queue.Full:
                pass
        
        

    

    def update_croppedimage(self, frame, recoreded_frame):
        """Updates the cropped image_label with a new cropped opencv image"""
        
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(310, 310, Qt.KeepAspectRatio)
        qt_cimg = QPixmap.fromImage(p)
        self.ui.CroppedVideoFeedLabel.setPixmap(qt_cimg)
        
        #recored the robots suroundings
        if self.croppedresult is not None:
            self.croppedresult.write(recoreded_frame)

    

    def croppedrecordfunction(self):
        if self.cap is not None:
            if self.ui.croppedrecordbutton.isChecked():
                self.ui.croppedrecordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
                file_path  = os.path.join(self.new_dir_path, self.date+".mp4")
                self.croppedresult = cv2.VideoWriter(
                    file_path,
                    cv2.VideoWriter_fourcc(*"mp4v"),
                    int(self.videofps),    
                    (200, 200), ) 
                #start recording magnetic field and tracking data
                self.start_data_record()
            
            else:
                self.ui.croppedrecordbutton.setText("Record")
                if self.croppedresult is not None:
                    self.croppedresult.release()
                    self.croppedresult = None
                    self.tbprint("End Record, Data Saved")
                #stop and save the data when the record is over.
                self.stop_data_record()
    
         
    

    
    def setFile(self):
        if self.videopath == 0:
            try:
                self.cap  = EasyPySpin.VideoCapture(0)
   
                self.cap.set(cv2.CAP_PROP_AUTO_WB, True)
                self.cap.set(cv2.CAP_PROP_FPS, 24)
                self.tbprint("Connected to FLIR Camera")

                if not self.cap.isOpened():
                    self.cap  = cv2.VideoCapture(0) 
                    self.tbprint("No EasyPySpin Camera Available")
            
            except Exception:
                self.cap  = cv2.VideoCapture(0) 
                self.tbprint("No EasyPySpin Camera Available")
                
                
            self.ui.pausebutton.hide()
            self.ui.leftbutton.hide()
            self.ui.rightbutton.hide()
            self.ui.frameslider.hide()
        else:
            self.cap  = cv2.VideoCapture(self.videopath)
            self.ui.pausebutton.show()
            self.ui.leftbutton.show()
            self.ui.rightbutton.show()
        
        self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
        self.tbprint("Width: {}  --  Height: {}  --  Fps: {}".format(self.video_width,self.video_height,self.videofps))

        self.aspectratio = (self.video_width / self.video_height)

        self.resize_widgets()        

        if self.videopath == 0:
            self.ui.robotsizeunitslabel.setText("um")
            self.ui.robotvelocityunitslabel.setText("um/s")
        else:
            self.ui.robotsizeunitslabel.setText("px")
            self.ui.robotvelocityunitslabel.setText("px/s")
            self.totalnumframes = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.tbprint("Total Frames: {} ".format(self.totalnumframes))
            self.ui.frameslider.setGeometry(QtCore.QRect(10, self.display_height+12, self.display_width, 20))
            self.ui.frameslider.setMaximum(self.totalnumframes)
            self.ui.frameslider.show()
        
        #if self.ui.recordbutton.isChecked():
            #self.recordfunction()

        #if not self.ui.startbutton.isChecked(): #clear the pixmap
        self.ui.VideoFeedLabel.setPixmap(QtGui.QPixmap())
        


    def selectFile(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.ReadOnly
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*);;Text Files (*.txt);;Python Files (*.py)", options=options)

        if file_path:
            self.videopath = file_path
            file_info = QtCore.QFileInfo(file_path)
            file_name = file_info.fileName()
            self.ui.choosevideobutton.setText(file_name)
            self.tbprint(file_name)
        else:
            self.videopath = 0
            self.ui.choosevideobutton.setText("Live")
            self.tbprint("Using Video Camera")
        
        self.setFile()
        
    
    
    def populate_serial_ports(self):
        ports = list_ports.comports()
        if len(ports) > 0:
            self.ui.arduino_portbox.clear()
            for port in ports:
                self.ui.arduino_portbox.addItem(port.device)
            self.arduino_port = port.device
        else:
            self.arduino_port = None

        

    def handle_port_change(self, selected_port):
        self.arduino_port = selected_port


    def start(self):

        if self.ui.startbutton.isChecked():
            #connect to arduino
            self.arduino = ArduinoHandler(self.tbprint, self.arduino_port)
            self.arduino.connect()
  

            if self.videopath is not None:

                self.frame_number = 0
                self.setFile()
                
                self.tracker = VideoThread(self)
                self.tracker.change_pixmap_signal.connect(self.update_image)
                self.tracker.cropped_frame_signal.connect(self.update_croppedimage)
                self.tracker.actions_signal.connect(self.update_actions)
                self.tracker.start()
               
                
                

                self.ui.startbutton.setText("Stop")
                self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
                self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
        
                
        else:
            self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
            self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
            
            if self.arduino is not None:
                self.arduino.close()
          


            if self.tracker is not None:
                self.ui.startbutton.setText("Start")
                self.tracker.stop()
                
                
                #reset control button
                self.control_status = False
                self.ui.controlbutton.setText("Control")
                self.tbprint("Control Off")
                self.ui.controlbutton.setChecked(False)

                #reset joystick button
                self.joystick_status = False
                self.ui.joystickbutton.setText("Joystick")
                self.tbprint("Joystick Off")
                self.ui.joystickbutton.setChecked(False)

                #reset mask button
                self.tracker.mask_flag = False
                self.ui.maskbutton.setText("Mask")
                self.ui.maskbutton.setChecked(False)

                #also reset pause button
                self.ui.pausebutton.setChecked(False)
                self.ui.pausebutton.setText("Pause")

                self.ui.pausebutton.hide()
                self.ui.leftbutton.hide()
                self.ui.rightbutton.hide()
            

                #zero arduino commands
                self.apply_actions(False)
                del self.tracker.robot_list[:]
                del self.magnetic_field_list[:]

                self.ui.applyacousticbutton.setChecked(False)
                self.ui.led.setStyleSheet("\n"
                "                background-color: rgb(255, 0, 0);\n"
                "                border-style: outset;\n"
                "                border-width: 3px;\n"
                "                border-radius: 12px;\n"
                "                border-color: rgb(255, 0, 0);\n"
                "         \n"
                "                padding: 6px;")

            

    def showmask(self):
        if self.tracker is not None:
            if self.ui.maskbutton.isChecked():
                self.ui.maskbutton.setText("Original")
                self.tracker.mask_flag = True
            else:
                self.ui.maskbutton.setText("Mask")
                self.tracker.mask_flag = False
    
    def showcroppedoriginal(self):
        if self.tracker is not None:
            if self.ui.croppedmasktoggle.isChecked():
                self.ui.croppedmasktoggle.setText("Mask")
                self.tracker.croppedmask_flag = False
            else:
                self.ui.croppedmasktoggle.setText("Original")
                self.tracker.croppedmask_flag = True


    def spinbox_alphachanged(self):
        self.ui.alphadial.setValue(self.ui.alphaspinBox.value())
    
    def dial_alphachanged(self):
        self.ui.alphaspinBox.setValue(self.ui.alphadial.value())

    def gradientcommand(self):
        self.gradient_status = int(self.ui.gradient_status_checkbox.isChecked())

    def equalfieldcommand(self):
        self.equal_field_status = int(self.ui.equal_field_checkbox.isChecked())

    def get_objective(self):
        if self.tracker is not None:
            self.tracker.objective = self.ui.objectivebox.value()

    def get_exposure(self):
        if self.tracker is not None:
            self.tracker.exposure = self.ui.exposurebox.value()
    
    def checkpush(self):
        if self.tracker is not None:
            self.tracker.pushstatus = self.ui.pushradio.isChecked()


    def checkorient(self):
        if self.tracker is not None:
            self.tracker.orientstatus = self.ui.orientradio.isChecked()

    def invertmaskcommand(self):
        if self.tracker is not None:
            self.ui.maskinvert_checkBox.setText("Invert Mask: " + str(self.ui.maskinvert_checkBox.isChecked()))
            self.tracker.maskinvert = self.ui.maskinvert_checkBox.isChecked()

    def pause(self):
        if self.videopath != 0:
            if self.ui.pausebutton.isChecked():
                self.tracker._play_flag = False
                self.ui.pausebutton.setText("Play")
              
            else:#play
                self.tracker._play_flag = True
                self.ui.pausebutton.setText("Pause")
                
    def frameright(self):
        if self.videopath != 0:
            self.tracker.framenum+=1
            self.ui.frameslider.setValue(self.tracker.framenum)
            self.ui.framelabel.setText("Frame:"+str(self.tracker.framenum))

    def frameleft(self):
        if self.videopath != 0:
            self.tracker.framenum-=1
            self.ui.frameslider.setValue(self.tracker.framenum)
            self.ui.framelabel.setText("Frame:"+str(self.tracker.framenum))

    
    
    
    def get_manual_bfieldbuttons(self):
        if self.ui.manualapplybutton.isChecked():
            self.manual_status = True
            self.ui.manualapplybutton.setText("Stop")
        else:
            self.ui.manualapplybutton.setText("Apply")
            self.apply_actions(False)
    


       
    def get_slider_vals(self):
        memory = self.ui.memorybox.value()
        RRTtreesize = self.ui.RRTtreesizebox.value()
        arrivalthresh = self.ui.arrivalthreshbox.value()
        magneticfreq = self.ui.magneticfrequencydial.value()
        gamma = self.ui.gammadial.value()
        psi = self.ui.psidial.value()
        #alpha = self.ui.alphaspinBox.value()
        
        robotlower = self.ui.robotmasklowerbox.value() 
        robotupper = self.ui.robotmaskupperbox.value()
        robotdilation = self.ui.robotmaskdilationbox.value() 
        robotmaskblur = self.ui.robotmaskblurbox.value()
        robotcrop_length = self.ui.robotcroplengthbox.value()
        
        celllower = self.ui.cellmasklowerbox.value() 
        cellupper = self.ui.cellmaskupperbox.value()
        celldilation = self.ui.cellmaskdilationbox.value() 
        cellmaskblur = self.ui.cellmaskblurbox.value()
        cellcrop_length = self.ui.cellcroplengthbox.value()
        

        if self.tracker is not None: 
            self.tracker.memory = memory
            self.tracker.RRTtreesize = RRTtreesize
            self.tracker.arrivalthresh = arrivalthresh
          
            self.tracker.robot_mask_lower = robotlower
            self.tracker.robot_mask_upper = robotupper
            self.tracker.robot_mask_dilation = robotdilation
            self.tracker.robot_mask_blur = robotmaskblur
            self.tracker.robot_crop_length = robotcrop_length
            
    
            self.tracker.cell_mask_lower = celllower
            self.tracker.cell_mask_upper = cellupper
            self.tracker.cell_mask_dilation = celldilation
            self.tracker.cell_mask_blur = cellmaskblur
            self.tracker.cell_crop_length = cellcrop_length
            

        self.ui.gammalabel.setText("Gamma: {}".format(gamma))
        self.ui.psilabel.setText("Psi: {}".format(psi))
        #self.ui.rollingfrequencylabel.setText("Freq: {}".format(magneticfreq))

         
        
    def resetparams(self):
        self.ui.robotmasklowerbox.setValue(0)
        self.ui.robotmaskupperbox.setValue(128)
        self.ui.robotmaskdilationbox.setValue(0)
        self.ui.robotmaskblurbox.setValue(0)
        self.ui.robotcroplengthbox.setValue(40)

        self.ui.cellmasklowerbox.setValue(0)
        self.ui.cellmaskupperbox.setValue(128)
        self.ui.cellmaskdilationbox.setValue(0)
        self.ui.cellmaskblurbox.setValue(0)
        self.ui.cellcroplengthbox.setValue(40)
        

    
        self.ui.memorybox.setValue(15)
        self.ui.RRTtreesizebox.setValue(25)
        self.ui.arrivalthreshbox.setValue(100)
        self.ui.gammadial.setSliderPosition(90)
        self.ui.psidial.setSliderPosition(90)
        self.ui.magneticfrequencydial.setValue(10)
        self.ui.acousticfreq_spinBox.setValue(1000000)
        self.ui.objectivebox.setValue(10)
        self.ui.exposurebox.setValue(5000)
        

    def resizeEvent(self, event):
        windowsize = event.size()
        self.window_width = windowsize.width()
        self.window_height = windowsize.height()
        self.resize_widgets()
 
    def resize_widgets(self):
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio

        self.display_width = int(self.display_height * self.aspectratio)

        self.ui.VideoFeedLabel.setGeometry(QtCore.QRect(10,  5,                       self.display_width,     self.display_height))
        self.ui.frameslider.setGeometry(QtCore.QRect(10,    self.display_height+12,   self.display_width,     self.framesliderheight))
        self.ui.plainTextEdit.setGeometry(QtCore.QRect(10,  self.display_height+20+self.framesliderheight,   self.display_width,     self.textheight))

        #self.ui.tabWidget.setGeometry(QtCore.QRect(12,  6,  260 ,     self.tabheight))

    def handle_zoom(self, frame):
        
        if self.zoomscale > 1:
            x = self.zoom_x
            y = self.zoom_y
            w = 300
            h = 300
            angle = 0
            
            # step 1: cropped a frame around the coord you wont to zoom into
            if y-w < 0 and x-h < 0:
                zoomedframe = frame[0:y+h , 0:x+w]
                cv2.rectangle(frame, (0, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = y
            elif x-w < 0:
                zoomedframe = frame[y-h:y+h , 0:x+w] 
                cv2.rectangle(frame, (0, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = h
            elif y-h < 0:
                zoomedframe = frame[0:y+h , x-w:x+w]
                cv2.rectangle(frame, (x-w, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = y
            else:
                zoomedframe = frame[y-h:y+h , x-w:x+w] 
                cv2.rectangle(frame, (x-w, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = h   
            
            # step 2: zoom into the zoomed frame a certain zoom amount
            rot_mat = cv2.getRotationMatrix2D((warpx,warpy), angle, self.zoomscale)
            zoomedframe = cv2.warpAffine(zoomedframe, rot_mat, zoomedframe.shape[1::-1], flags=cv2.INTER_LINEAR)

            #step 3: replace the original cropped frame with the new zoomed in cropped frame
            if y-h < 0 and x-w < 0:
                frame[0:y+h , 0:x+w] =  zoomedframe
            elif x-w < 0:
                frame[y-h:y+h , 0:x+w] =  zoomedframe
            elif y-h < 0:
                frame[0:y+h , x-w:x+w] =  zoomedframe
            else:
                frame[y-h:y+h , x-w:x+w] =  zoomedframe


        
        return frame

    def closeEvent(self, event):
        """
        called when x button is pressed
        """
        
        if self.tracker is not None:
            self.tracker.stop()
        #self.recorder.stop()
        
        self.simulator.stop()
        self.apply_actions(False)
        if self.arduino is not None:
            self.arduino.close()
